"""
Logbook Flask server — Port Cares In-Kind Donations

POST /append  — receives staged forms, appends to correct XLSX tab in date order
POST /ocr     — receives base64 image, returns extracted fields via Claude Vision
GET  /status  — health check
"""

from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import datetime
import json
import anthropic
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
client = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY'))

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# Column order matches Port Cares format
COLUMNS = [
    ('A', 'DATE',           'date'),
    ('B', 'SOURCE',         'donor'),
    ('C', 'PERISHABLE',     None),         # computed
    ('D', 'NON-PERISHABLE', 'nonPerishable'),
    ('E', 'MEAT',           'meat'),
    ('F', 'DAIRY',          'dairy'),
    ('G', 'EGGS',           None),         # always blank
    ('H', 'PRODUCE',        'produce'),
    ('I', 'MONTHLY TOTAL',  None),         # computed: C+D
    ('J', 'YTD',            None),         # blank, Port Cares fills manually
]

PERISHABLE_FIELDS = ['produce', 'dairy', 'meat', 'bakedGoods']

HEADER_FILL = PatternFill(start_color='185FA5', end_color='185FA5', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
BOLD_FONT   = Font(bold=True)

INSTRUCTIONS = (
    'INSTRUCTIONS: Enter as per weigh-in logs by scale in WH. '
    'Always enter value under perishable or non-perishable AND under '
    'appropriate category i.e. "produce" or "meat".'
)

MONTH_NAMES = {
    1:'JAN',2:'FEB',3:'MAR',4:'APR',5:'MAY',6:'JUN',
    7:'JUL',8:'AUG',9:'SEP',10:'OCT',11:'NOV',12:'DEC',
}

TABLE_NAME_PREFIX = 'InKind'

OCR_PROMPT = """You are extracting data from a PortCares In-Kind Donations form.

The form has these fields:
- Date Received (mm/dd/yyyy)
- Donor/Event (free text, e.g. "SDM", "Port BIC", "Anon")
- Weight (total weight in lbs)
- Categories with weights written next to the circled ones:
  Non-Perishable, Produce, Dairy, Meat, Baked Goods, Pet Food, Toys, Hygiene, School Supplies
- Other (free text, usually blank)
- Contact info (often blank): Name/Business, Address, Email, Phone, New Donor (Yes/No)

Extract all visible values and return ONLY a JSON object with these exact keys:
{
  "date": "",
  "donor": "",
  "weight": "",
  "nonPerishable": "",
  "produce": "",
  "dairy": "",
  "meat": "",
  "bakedGoods": "",
  "petFood": "",
  "toys": "",
  "hygiene": "",
  "schoolSupplies": "",
  "other": "",
  "contactName": "",
  "contactAddress": "",
  "contactEmail": "",
  "contactPhone": "",
  "newDonor": ""
}

Rules:
- Only include values you can clearly read. Leave blank if uncertain.
- For categories, only include a value if that category is circled/selected AND has a number written next to it.
- Weight values should be numbers only (no "lbs" or units).
- Date should be in mm/dd/yyyy format.
- newDonor should be "yes", "no", or "" if not filled.
- If a value appears crossed out with a new value written nearby, use the new value only.
- If you cannot determine which value is correct due to corrections or unclear writing, make your best guess — the reviewer will verify.
- Return ONLY the JSON object, no other text.
"""


def get_xlsx_path(year: str) -> str:
    return os.path.join(DATA_DIR, f'Port Cares In {year}.xlsx')


def get_or_create_workbook(year: str):
    path = get_xlsx_path(year)
    if os.path.exists(path):
        return load_workbook(path), path
    wb = Workbook()
    wb.remove(wb.active)
    return wb, path


def setup_sheet(ws):
    """
    Row 1: Instructions (merged, italic)
    Row 2: MONTHLY TOTAL with SUM formulas referencing the table
    Row 3: Table headers (blue)
    Row 4+: Data (managed by table)
    """
    table_name = f"{TABLE_NAME_PREFIX}_{ws.title}"

    # Row 1: Instructions
    ws.merge_cells('A1:J1')
    ws['A1'] = INSTRUCTIONS
    ws['A1'].font = Font(italic=True, size=9)
    ws['A1'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Row 2: MONTHLY TOTAL with table-reference SUM formulas
    ws['A2'] = 'MONTHLY TOTAL'
    ws['A2'].font = BOLD_FONT
    # SUM formulas for numeric columns — reference table by name
    for col_letter, header, _ in COLUMNS:
        if col_letter in ('A', 'B'):
            continue
        col_idx = ord(col_letter) - ord('A') + 1
        ws.cell(row=2, column=col_idx).value = f'=SUM({table_name}[{header}])'

    # Row 3: Table headers (styled blue — table will override but we style anyway)
    for col_letter, header, _ in COLUMNS:
        col_idx = ord(col_letter) - ord('A') + 1
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    widths = {'A':10,'B':20,'C':14,'D':16,'E':10,'F':10,'G':8,'H':10,'I':14,'J':10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Create table starting at row 3, initial range covers header + 1 empty row
    # (openpyxl requires at least 2 rows for a table)
    tab = Table(displayName=table_name, ref='A3:J4')
    style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def get_table(ws):
    """Return the InKind table for this sheet, or None."""
    for tbl in ws.tables.values():
        if tbl.displayName.startswith(TABLE_NAME_PREFIX):
            return tbl
    return None


def extend_table(ws, new_last_row):
    """Extend the table ref to include new_last_row."""
    tbl = get_table(ws)
    if tbl:
        # Keep column range A:J, extend rows
        tbl.ref = f'A3:J{new_last_row}'


def parse_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.datetime.strptime(date_str.strip(), '%m/%d/%Y').date()
    except Exception:
        return None


def compute_row_values(form):
    def num(field):
        try:
            v = form.get(field, '')
            return float(v) if v else 0
        except (ValueError, TypeError):
            return 0

    perishable    = sum(num(f) for f in PERISHABLE_FIELDS)
    non_perishable = num('nonPerishable')
    meat          = num('meat')
    dairy         = num('dairy')
    produce       = num('produce')
    monthly_total = perishable + non_perishable

    return {
        'A': parse_date(form.get('date', '')),
        'B': form.get('donor') or None,
        'C': perishable      if perishable > 0      else None,
        'D': non_perishable  if non_perishable > 0  else None,
        'E': meat            if meat > 0            else None,
        'F': dairy           if dairy > 0           else None,
        'G': None,
        'H': produce         if produce > 0         else None,
        'I': monthly_total   if monthly_total > 0   else None,
        'J': None,
    }


def find_insert_row(ws, target_date):
    """Find correct insertion row to maintain date order. Data starts at row 4."""
    last_data_row = 3
    for row in range(4, ws.max_row + 1):
        if any(ws.cell(row=row, column=col).value is not None for col in range(1, 11)):
            last_data_row = row

    if target_date is None:
        return last_data_row + 1

    insert_after = 3
    last_seen_date = None

    for row in range(4, last_data_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val is not None:
            row_date = cell_val.date() if hasattr(cell_val, 'date') else parse_date(str(cell_val))
            if row_date:
                last_seen_date = row_date
            if last_seen_date and last_seen_date <= target_date:
                insert_after = row
            elif last_seen_date and last_seen_date > target_date:
                break
        else:
            # Blank date row — belongs to last_seen_date group
            if last_seen_date and last_seen_date <= target_date:
                insert_after = row

    return insert_after + 1


def insert_form_row(ws, row_num, values):
    ws.insert_rows(row_num)
    for col_letter, value in values.items():
        col_idx = ord(col_letter) - ord('A') + 1
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = value
        if col_letter == 'A' and isinstance(value, datetime.date):
            cell.number_format = 'D-MMM'


@app.route('/append', methods=['POST'])
def append_forms():
    data = request.get_json()
    if not data or 'forms' not in data:
        return jsonify({'ok': False, 'error': 'Missing forms'}), 400

    forms = data['forms']
    if not forms:
        return jsonify({'ok': False, 'error': 'No forms provided'}), 400

    rows_added = 0
    new_files = []

    # Group by year and month
    by_year_month = {}
    for form in forms:
        date = parse_date(form.get('date', ''))
        year = str(date.year) if date else str(datetime.datetime.now().year)
        month_num = date.month if date else datetime.datetime.now().month
        by_year_month.setdefault(year, {}).setdefault(month_num, []).append((form, date))

    for year, by_month in by_year_month.items():
        file_existed = os.path.exists(get_xlsx_path(year))
        if not file_existed:
            new_files.append(year)
        wb, path = get_or_create_workbook(year)

        for month_num, form_date_pairs in by_month.items():
            month_name = MONTH_NAMES[month_num]

            if month_name in wb.sheetnames:
                ws = wb[month_name]
            else:
                ws = wb.create_sheet(month_name)
                setup_sheet(ws)

            form_date_pairs.sort(key=lambda x: x[1] or datetime.date.min)

            for form, date in form_date_pairs:
                values = compute_row_values(form)
                insert_row = find_insert_row(ws, date)

                above_val = ws.cell(row=insert_row - 1, column=1).value
                if above_val is not None:
                    above_date = above_val.date() if hasattr(above_val, 'date') else parse_date(str(above_val))
                    if date and above_date == date:
                        values['A'] = None

                insert_form_row(ws, insert_row, values)
                rows_added += 1

            extend_table(ws, ws.max_row)
        wb.save(path)

    return jsonify({'ok': True, 'rows_added': rows_added, 'new_files': new_files})

@app.route('/ocr', methods=['POST'])
def ocr():
    data = request.get_json()
    if not data or 'image' not in data:
        return jsonify({'ok': False, 'error': 'Missing image'}), 400

    image_data = data['image']
    if ',' in image_data:
        image_data = image_data.split(',')[1]

    try:
        message = client.messages.create(
            model='claude-haiku-4-5',
            max_tokens=1024,
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'image', 'source': {'type': 'base64', 'media_type': 'image/jpeg', 'data': image_data}},
                    {'type': 'text', 'text': OCR_PROMPT},
                ],
            }],
        )
        raw = message.content[0].text.strip()
        if raw.startswith('```'):
            raw = raw.split('```')[1]
            if raw.startswith('json'):
                raw = raw[4:]
        fields = json.loads(raw.strip())
        return jsonify({'ok': True, 'fields': fields})

    except Exception as e:
        print(f'OCR error: {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/status', methods=['GET'])
def status():
    files = os.listdir(DATA_DIR) if os.path.exists(DATA_DIR) else []
    return jsonify({'ok': True, 'files': files})


@app.route('/', methods=['GET'])
def index():
    return 'Logbook server running.'


if __name__ == '__main__':
    print(f'Logbook server starting on http://0.0.0.0:5000')
    print(f'Data directory: {DATA_DIR}')
    app.run(host='0.0.0.0', port=5000, debug=True)